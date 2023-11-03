import openpyxl
import hashlib
import serial
from pyfiglet import Figlet
import mysql.connector
from getpass import getpass
import bcrypt
import json

# MySQL Database Configuration
db = mysql.connector.connect(
    host="localhost",
    user="root",
    password="Password@123",
    database="mydb"
)
cursor = db.cursor()

def create_table_if_not_exists():
    create_table_query = f"""
    CREATE TABLE IF NOT EXISTS users (
        id INT AUTO_INCREMENT PRIMARY KEY,
        username VARCHAR(29) UNIQUE,
        password VARCHAR(29)
    )
    """
    cursor.execute(create_table_query)

def check_data():
    select_query = f"SELECT * FROM users"
    cursor.execute(select_query)
    return cursor.fetchall()

# Register new user account
def create_account(username, password):
    create_table_if_not_exists()
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
    cursor.execute("INSERT INTO users (username, password) VALUES (%s, %s)", (username, hashed_password))
    db.commit()
    return "Account created successfully"

def check_user(username):
    cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
    user = cursor.fetchone()
    if user:
        return True
    else:
        return False

def change_password(username, password):
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
    cursor.execute("UPDATE users SET password = %s WHERE username = %s", (hashed_password, username))
    print("Password updated")



# Login 
def login(username, password):
    cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
    user = cursor.fetchone()
    if user and bcrypt.checkpw(password.encode('utf-8'), user[2].encode('utf-8')):
        print("Login successful!")
        return True
    else:
        print("Login failed. Please check your credentials.")
        return False

# Function to check if a username and password match the Excel data
# def login(username, password):
#     # Load the Excel file
#     workbook = openpyxl.load_workbook('user_data.xlsx')
#     sheet = workbook.active

#     # Convert the entered password to MD5
#     password_md5 = hashlib.md5(password.encode()).hexdigest()

#     # Iterate through rows to check for a matching username
#     for row in sheet.iter_rows(values_only=True):
#         if row[0] == username and row[1] == password_md5:
#             return True

#     return False


# Function to create a new account and append it to the Excel file
# def create_account(username, password):
#     # Load the Excel file
#     workbook = openpyxl.load_workbook('user_data.xlsx')
#     sheet = workbook.active

#     # Check if the username already exists
#     for row in sheet.iter_rows(values_only=True):
#         if row[0] == username:
#             return "Username already exists. Please choose a different username."

#     # Convert the password to MD5 before saving it
#     password_md5 = hashlib.md5(password.encode()).hexdigest()

#     # If the username is not already in use, append the new account
#     sheet.append([username, password_md5])
#     workbook.save('user_data.xlsx')
#     return "Account created successfully."


def send_message(user, message):
    data = {
        'user': user,
        'message': message.decode('utf-8')
    }
    json_string = json.dumps(data)
    return json_string

# Function to open a chat session with Arduino
def chat_with_arduino(ser, username):
    print("Chatting with Arduino. Type your messages and press Enter to send. Type 'QUIT!' to exit. ")
    while True:
        user_message = input("You: ")
        user_message_map = send_message(username, user_message.encode())
        ser.write(user_message_map.encode('utf-8'))
        # arduino_response = ser.readline().decode().strip()
        # print("Arduino:", arduino_response)
        if user_message == "QUIT!":
            break


serport = serial.Serial('COM7', 9600) 

# Main program
def main():
    # Create a Figlet object for ASCII art text
    custom_figlet = Figlet(font='slant')

    # Display the ASCII art banner
    print(custom_figlet.renderText("Welcome to LORA Chat Platform"))


    while True:
        print("1. Login")
        print("2. Create new account")
        print("3. Change Password")
        choice = input("Enter your choice (1 or 2): ")

        if choice == '1':
            usernameinput = input("Enter your username: ")
            passwordinput = getpass("Enter your password: ")
            if login(usernameinput, passwordinput):
                print("You are logged in.")
                chat_with_arduino(serport, usernameinput)
            else:
                print("Username or password is wrong.")

        elif choice == '2':
            newusername = input("Enter a username: ")
            newpassword = getpass("Enter a password: ")
            result = create_account(newusername, newpassword)
            print(result)

        elif choice == '3':
            username = input("Enter a username: ")
            if (check_user(username)):
                newpassword = getpass("Enter the new password: ")
                newpasswordAgain = getpass("Enter the new password again: ")
                if newpassword == newpasswordAgain:
                    change_password(username, newpasswordAgain)
                else:
                    print("Passwords do not match")
            else:
                print("User does not exist")


        else:
            print("Invalid choice. Please select 1 or 2.")


if __name__ == "__main__":
    main()
